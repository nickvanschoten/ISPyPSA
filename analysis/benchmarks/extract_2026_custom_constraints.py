"""Extract the 21 REZ group / transmission-limit constraints from the Draft 2026
ISP (v7.5) workbook's "Build limits - REZs" sheet.

These constraints can't be read by `isp-workbook-parser` (irregular merged-cell
layout), so ISPyPSA ships them as a manually-extracted table. This script makes
the extraction reproducible rather than hand-transcribed: it reads the three
constraint sub-tables, resolves each constraint's row-span from the merged col-D
ranges, and pulls the coefficient out of the col-B term text.

Output is the *raw* extraction (one row per (constraint, term)) with the
coefficient separated from the term label. Binding term labels to PyPSA
component names is a separate step (see notes in the v7.5 migration writeup).

Run: PYTHONPATH=. uv run python analysis/benchmarks/extract_2026_custom_constraints.py
"""

import re
from pathlib import Path

import openpyxl
import pandas as pd

WORKBOOK = Path("iasr inputs/Draft 2026 ISP Inputs and Assumptions workbook.xlsx")
SHEET = "Build limits - REZs"
OUT = Path("analysis/benchmarks/_2026_constraints_raw.csv")

# Constraint rows live in three sub-tables on the sheet (header rows excluded).
CONSTRAINT_ROW_RANGES = [(136, 211), (217, 248), (254, 266)]

_COEF_TIMES = re.compile(r"^\s*([+-]?\d*\.?\d+)\s*\*\s*(.+?)\s*$")  # "0.14 * WD"
_LEADING_MINUS = re.compile(r"^\s*-\s*(.+?)\s*$")  # " - CQ-NQ" => -1 * CQ-NQ


def _parse_coefficient(term_text: str) -> tuple[float, str]:
    """Split a col-B term cell into (coefficient, term label).

    "0.14 * WD" -> (0.14, "WD");  " - CQ-NQ" -> (-1.0, "CQ-NQ");  "Q7" -> (1.0, "Q7").
    """
    m = _COEF_TIMES.match(term_text)
    if m:
        return float(m.group(1)), m.group(2).strip()
    m = _LEADING_MINUS.match(term_text)
    if m:
        return -1.0, m.group(1).strip()
    return 1.0, term_text.strip()


def _resolve_merged(ws) -> dict:
    """Map every (row, col) inside a merged range to that range's anchor value."""
    resolved = {}
    for rng in ws.merged_cells.ranges:
        anchor = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                resolved[(r, c)] = anchor
    return resolved


def _cell(ws, resolved, r, c):
    return resolved.get((r, c), ws.cell(row=r, column=c).value)


def _extract_rows(ws, resolved) -> list[dict]:
    rows = []
    for lo, hi in CONSTRAINT_ROW_RANGES:
        for r in range(lo, hi + 1):
            term_text = _cell(ws, resolved, r, 2)  # col B
            constraint_id = _cell(ws, resolved, r, 4)  # col D (merged per constraint)
            if term_text is None or constraint_id == "Term":
                continue
            coefficient, term = _parse_coefficient(str(term_text))
            summer_peak = _cell(ws, resolved, r, 5)  # col E
            import_limit = _cell(ws, resolved, r, 8)  # col H
            rows.append(
                {
                    "constraint_id": constraint_id,
                    "term": term,
                    "coefficient": coefficient,
                    "rhs_summer_peak": summer_peak,
                    "import_limit": import_limit,
                    "source_row": r,
                }
            )
    return rows


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    resolved = _resolve_merged(ws)
    df = pd.DataFrame(_extract_rows(ws, resolved))
    df.to_csv(OUT, index=False)
    n_constraints = df["constraint_id"].nunique()
    print(f"Extracted {len(df)} terms across {n_constraints} constraints -> {OUT}")
    for cid, grp in df.groupby("constraint_id", sort=False):
        rhs = grp["rhs_summer_peak"].iloc[0]
        imp = grp["import_limit"].iloc[0]
        bound = f"RHS<={rhs}" if rhs not in (None, "N/A") else f"import<={imp}"
        terms = ", ".join(f"{c:+g}*{t}" for c, t in zip(grp["coefficient"], grp["term"]))
        print(f"  {cid} [{bound}] ({len(grp)} terms): {terms}")


if __name__ == "__main__":
    main()
